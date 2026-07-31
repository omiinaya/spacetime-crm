"""CSV / XLSX / JSON export and import routes.

The export endpoint (`/api/export/{entity}`) serves CSV by default and
XLSX/JSON via the ``format`` query parameter. The import endpoints
(`/api/import/customers`, `/api/import/products`) accept all three formats
and auto-detect the file type from its content, reusing the same row-parsing
and field-validation logic for every format.
"""

from __future__ import annotations

import csv
import io
import json
import time

from fastapi import APIRouter, Depends, File, HTTPException, Query, Response, UploadFile
from helpers import (
    CUSTOMER_SENSITIVE_FIELDS,
    _call,
    _sql,
    _sqlesc,
    require_role,
)

router = APIRouter()

ENTITY_TABLE_MAP = {
    "customers": "customer",
    "tickets": "ticket",
    "invoices": "invoices",
    "payments": "payment",
    "appointments": "appointment",
    "products": "products",
    "estimates": "estimates",
    "purchase_orders": "purchase_order",
    "tax_rates": "tax_rate",
    "users": "user",
}

EXPORT_FORMATS = ("csv", "xlsx", "json")

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# ── Shared cell / row helpers (CSV, XLSX and JSON all reduce to row dicts) ──


def _cell_str(value) -> str:
    """Normalize any cell value (str/num/bool/None) to a stripped string."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value).strip()


def _cell_default(value, default):
    """Return ``value`` when it is non-empty, otherwise ``default``."""
    if _cell_str(value) == "":
        return default
    return value


def _parse_bool(value) -> bool:
    """Parse a boolean from CSV strings, XLSX/JSON bools or numbers."""
    if isinstance(value, bool):
        return value
    return _cell_str(value).lower() in ("true", "1", "yes")


def _rows_from_csv(text: str) -> tuple[list[dict], list[str]]:
    reader = csv.DictReader(io.StringIO(text))
    return list(reader), list(reader.fieldnames or [])


def _rows_from_xlsx(content: bytes) -> tuple[list[dict], list[str]]:
    """Read an XLSX workbook: first row is the header, remaining rows are data."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    assert ws is not None
    it = ws.iter_rows(values_only=True)
    try:
        header = next(it)
    except StopIteration:
        return [], []
    headers = [_cell_str(h) for h in header] if header else []
    rows = []
    for values in it:
        if values is None:
            continue
        row = {}
        for i, h in enumerate(headers):
            row[h] = values[i] if i < len(values) else None
        rows.append(row)
    return rows, headers


def _rows_from_json(text: str) -> tuple[list[dict], list[str]]:
    data = json.loads(text)
    if not isinstance(data, list) or any(not isinstance(r, dict) for r in data):
        raise HTTPException(400, "JSON import must be an array of objects")
    fieldnames: list[str] = []
    for r in data:
        for k in r:
            if k not in fieldnames:
                fieldnames.append(k)
    return data, fieldnames


def _payload_rows(content: bytes) -> tuple[list[dict], list[str], str]:
    """Detect the file format and reduce the payload to row dicts.

    Returns ``(rows, fieldnames, format_name)``. Detection is by content:
    XLSX files start with the ZIP magic ``PK``, JSON starts with ``{`` or
    ``[``, anything else is treated as CSV (including a UTF-8 BOM).
    """
    if content[:4] == b"PK\x03\x04":
        return (*_rows_from_xlsx(content), "xlsx")
    if content.lstrip()[:1] in (b"{", b"["):
        return (*_rows_from_json(content.decode("utf-8-sig")), "json")
    return (*_rows_from_csv(content.decode("utf-8-sig")), "csv")


async def _import_customers_rows(rows: list[dict], user: dict) -> dict:
    """Shared customer import — used by CSV, XLSX and JSON.

    Required: first_name. If an ``id`` column is present with a value, uses
    the ``import_customer`` reducer to preserve IDs; otherwise creates new
    customers via ``create_customer``.
    """
    now_ms = int(time.time() * 1000)
    count = 0
    errors = []
    for i, row in enumerate(rows, start=2):
        try:
            fn = _cell_str(row.get("first_name"))
            ln = _cell_str(row.get("last_name"))
            email = _cell_str(row.get("email"))
            phone = _cell_str(row.get("phone"))
            mobile = _cell_str(row.get("mobile"))
            company = _cell_str(row.get("company"))
            # Accept both spellings: DB/export uses address_line_1, the legacy
            # CSV import format uses address_line1. Explicit value wins.
            addr1 = _cell_str(row.get("address_line1") or row.get("address_line_1"))
            addr2 = _cell_str(row.get("address_line2") or row.get("address_line_2"))
            city = _cell_str(row.get("city"))
            state = _cell_str(row.get("state"))
            zipc = _cell_str(row.get("zip"))
            notes = _cell_str(row.get("notes"))
            tags = _cell_str(row.get("tags"))

            cid = _cell_str(row.get("id"))
            if cid:
                created_at = int(row.get("created_at") or now_ms)
                updated_at = int(row.get("updated_at") or now_ms)
                await _call(
                    "import_customer",
                    [
                        user["tenant_id"],
                        cid,
                        fn,
                        ln,
                        email,
                        phone,
                        mobile,
                        addr1,
                        addr2,
                        city,
                        state,
                        zipc,
                        company,
                        notes,
                        tags,
                        created_at,
                        updated_at,
                    ],
                )
            else:
                await _call("create_customer", [user["tenant_id"], fn, ln, email, phone])
            count += 1
        except Exception as e:
            errors.append(f"Row {i}: {e}")
    return {"imported": count, "errors": errors}


async def _import_products_rows(rows: list[dict], user: dict) -> dict:
    """Shared product import — used by CSV, XLSX and JSON.

    Required: name. If an ``id`` column is present with a value, uses the
    ``import_product`` reducer to preserve IDs; otherwise creates new
    products via ``create_product``.
    """
    now_ms = int(time.time() * 1000)
    count = 0
    errors = []
    for i, row in enumerate(rows, start=2):
        try:
            name = _cell_str(row.get("name"))
            sku = _cell_str(row.get("sku"))
            barcode = _cell_str(row.get("barcode"))
            desc = _cell_str(row.get("description"))
            category = _cell_str(row.get("category"))
            price = float(row.get("price") or 0)
            cost = float(row.get("cost") or 0)
            qoh = float(row.get("quantity_on_hand") or 0)
            qc = float(row.get("quantity_committed") or 0)
            min_stock = float(row.get("min_stock") or 0)
            reorder_quantity = float(row.get("reorder_quantity") or 0)
            location = _cell_str(row.get("location"))
            active = _parse_bool(_cell_default(row.get("active"), "true"))

            pid = _cell_str(row.get("id"))
            if pid:
                created_at = int(row.get("created_at") or now_ms)
                updated_at = int(row.get("updated_at") or now_ms)
                await _call(
                    "import_product",
                    [
                        user["tenant_id"],
                        pid,
                        name,
                        sku,
                        barcode,
                        desc,
                        category,
                        price,
                        cost,
                        qoh,
                        qc,
                        min_stock,
                        reorder_quantity,
                        location,
                        active,
                        created_at,
                        updated_at,
                    ],
                )
            else:
                await _call(
                    "create_product",
                    [
                        user["tenant_id"],
                        name,
                        sku,
                        barcode,
                        desc,
                        category,
                        price,
                        cost,
                        qoh,
                        min_stock,
                        reorder_quantity,
                        location,
                    ],
                )
            count += 1
        except Exception as e:
            errors.append(f"Row {i}: {e}")
    return {"imported": count, "errors": errors}


# ── EXPORT ──


@router.get("/api/export/{entity}")
async def export_entity(
    entity: str,
    format: str = Query("csv", description="csv | xlsx | json"),
    user: dict = Depends(require_role("admin", "tech", "front_desk")),
):
    """Export all records of an entity type as CSV (default), XLSX or JSON.

    Scoped to the caller's tenant; sensitive fields (e.g. customer password
    hashes) are stripped. Downloads as an attachment.
    """
    table = ENTITY_TABLE_MAP.get(entity)
    if not table:
        raise HTTPException(400, f"Unknown entity: {entity}. Valid: {', '.join(ENTITY_TABLE_MAP)}")

    fmt = (format or "csv").lower()
    if fmt not in EXPORT_FORMATS:
        raise HTTPException(
            400, f"Unsupported format: {format}. Valid: {', '.join(EXPORT_FORMATS)}"
        )

    rows = await _sql(f"SELECT * FROM {table} WHERE tenant_id = '{_sqlesc(user['tenant_id'])}'")
    if entity == "customers":
        rows = [{k: v for k, v in r.items() if k not in CUSTOMER_SENSITIVE_FIELDS} for r in rows]

    disposition = f'attachment; filename="{entity}.{fmt}"'

    if fmt == "csv":
        output = io.StringIO()
        if rows:
            writer = csv.DictWriter(output, fieldnames=rows[0].keys())
            writer.writeheader()
            writer.writerows(rows)
        return Response(
            content=output.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": disposition},
        )

    if fmt == "json":
        return Response(
            content=json.dumps(rows, indent=2, default=str),
            media_type="application/json",
            headers={"Content-Disposition": disposition},
        )

    # fmt == "xlsx"
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = entity[:31] or "data"
    if rows:
        headers = list(rows[0].keys())
        ws.append(headers)
        for r in rows:
            ws.append([r.get(h) for h in headers])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type=XLSX_MIME,
        headers={"Content-Disposition": disposition},
    )


# ── IMPORT ──


@router.post("/api/import/customers")
async def import_customers(
    file: UploadFile = File(...), user: dict = Depends(require_role("admin"))
):
    """Import customers from CSV, XLSX or JSON (format auto-detected).

    Required: first_name. Optional: email, phone, etc. If id column provided,
    uses import_customer reducer to preserve IDs.
    """
    content = await file.read()
    rows, fieldnames, _fmt = _payload_rows(content)

    if "first_name" not in fieldnames:
        raise HTTPException(400, "Import file must contain a 'first_name' column")

    result = await _import_customers_rows(rows, user)
    result["file"] = file.filename
    return result


@router.post("/api/import/products")
async def import_products(
    file: UploadFile = File(...), user: dict = Depends(require_role("admin"))
):
    """Import products from CSV, XLSX or JSON (format auto-detected).

    Required: name. Optional: sku, barcode, description, category, price,
    cost, quantity_on_hand, quantity_committed, min_stock, location, active.
    If id column provided, uses import_product reducer to preserve IDs.
    """
    content = await file.read()
    rows, fieldnames, _fmt = _payload_rows(content)

    if "name" not in fieldnames:
        raise HTTPException(400, "Import file must contain a 'name' column")

    result = await _import_products_rows(rows, user)
    result["file"] = file.filename
    return result
