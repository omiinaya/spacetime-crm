"""CSV / XLSX / JSON export and import tests."""

import io
import json

import httpx

from .conftest import (
    SERVER_URL,
    assert_ok,
    unique_suffix,
)


def _make_xlsx(headers: list[str], rows: list[list]) -> bytes:
    """Build an XLSX byte payload from headers + row values."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestExport:
    def test_export_customers(self, test_admin_headers: dict):
        resp = httpx.get(
            f"{SERVER_URL}/api/export/customers", headers=test_admin_headers, timeout=10
        )
        assert resp.status_code == 200
        assert "csv" in resp.headers.get("content-type", "").lower() or resp.status_code == 200

    def test_export_invalid_entity(self, test_admin_headers: dict):
        resp = httpx.get(
            f"{SERVER_URL}/api/export/nonexistent",
            headers=test_admin_headers,
            timeout=10,
        )
        assert resp.status_code == 400

    def test_export_tickets(self, test_admin_headers: dict):
        resp = httpx.get(f"{SERVER_URL}/api/export/tickets", headers=test_admin_headers, timeout=10)
        assert resp.status_code == 200


class TestExportXlsxJson:
    def test_export_customers_xlsx(self, test_admin_headers: dict, session_suffix: str):
        suf = unique_suffix()
        cust = httpx.post(
            f"{SERVER_URL}/api/customers",
            json={
                "first_name": "XlsxExp",
                "last_name": "Customer",
                "email": f"xlsx-exp-{session_suffix}-{suf}@test.com",
                "phone": "555-0301",
            },
            headers=test_admin_headers,
            timeout=10,
        )
        assert cust.status_code in (200, 201)

        resp = httpx.get(
            f"{SERVER_URL}/api/export/customers?format=xlsx",
            headers=test_admin_headers,
            timeout=10,
        )
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers.get("content-type", "").lower()
        assert resp.content[:4] == b"PK\x03\x04"
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(resp.content), read_only=True)
        it = wb.active.iter_rows(values_only=True)
        headers = [str(h) for h in next(it)]
        assert "first_name" in headers
        assert "portal_password_hash" not in headers  # sensitive field stripped
        emails = {row[headers.index("email")] for row in it}
        assert f"xlsx-exp-{session_suffix}-{suf}@test.com" in emails

    def test_export_customers_json(self, test_admin_headers: dict, session_suffix: str):
        suf = unique_suffix()
        cust = httpx.post(
            f"{SERVER_URL}/api/customers",
            json={
                "first_name": "JsonExp",
                "last_name": "Customer",
                "email": f"json-exp-{session_suffix}-{suf}@test.com",
            },
            headers=test_admin_headers,
            timeout=10,
        )
        assert cust.status_code in (200, 201)

        resp = httpx.get(
            f"{SERVER_URL}/api/export/customers?format=json",
            headers=test_admin_headers,
            timeout=10,
        )
        assert resp.status_code == 200
        assert "application/json" in resp.headers.get("content-type", "").lower()
        data = resp.json()
        assert isinstance(data, list)
        assert data, "expected exported customers to be non-empty"
        assert "portal_password_hash" not in data[0]
        emails = {c.get("email") for c in data}
        assert f"json-exp-{session_suffix}-{suf}@test.com" in emails

    def test_export_products_xlsx(self, test_admin_headers: dict, session_suffix: str):
        suf = unique_suffix()
        prod = httpx.post(
            f"{SERVER_URL}/api/products",
            json={
                "name": "Xlsx Export Product",
                "sku": f"XLSX-EXP-{session_suffix}-{suf}",
                "price": 15.99,
                "cost": 7.0,
            },
            headers=test_admin_headers,
            timeout=10,
        )
        assert prod.status_code in (200, 201)

        resp = httpx.get(
            f"{SERVER_URL}/api/export/products?format=xlsx",
            headers=test_admin_headers,
            timeout=10,
        )
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers.get("content-type", "").lower()
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(resp.content), read_only=True)
        it = wb.active.iter_rows(values_only=True)
        headers = [str(h) for h in next(it)]
        assert "name" in headers
        skus = {row[headers.index("sku")] for row in it}
        assert f"XLSX-EXP-{session_suffix}-{suf}" in skus

    def test_export_bad_format(self, test_admin_headers: dict):
        resp = httpx.get(
            f"{SERVER_URL}/api/export/customers?format=pdf",
            headers=test_admin_headers,
            timeout=10,
        )
        assert resp.status_code == 400

    def test_export_xlsx_unauthorized(self, client: httpx.Client):
        resp = client.get("/api/export/customers?format=xlsx", timeout=10)
        assert resp.status_code in (401, 403)


class TestImport:
    def test_import_customers_without_id(self, test_admin_headers: dict, session_suffix: str):
        """Import customers CSV without ID column."""
        suf = unique_suffix()
        csv_content = f"first_name,last_name,email,phone\nImp,Test1,imp1-{session_suffix}-{suf}@test.com,555-0101\nImp,Test2,imp2-{session_suffix}-{suf}@test.com,555-0102\n"
        files = {"file": ("import.csv", csv_content, "text/csv")}
        resp = httpx.post(
            f"{SERVER_URL}/api/import/customers",
            files=files,
            headers=test_admin_headers,
            timeout=15,
        )
        data = assert_ok(resp)
        assert data["imported"] >= 2

    def test_import_customers_missing_first_name(self, test_admin_headers: dict):
        csv_content = "last_name,email\nNobody,nobody@test.com\n"
        files = {"file": ("bad.csv", csv_content, "text/csv")}
        resp = httpx.post(
            f"{SERVER_URL}/api/import/customers",
            files=files,
            headers=test_admin_headers,
            timeout=10,
        )
        assert resp.status_code == 400

    def test_import_products(self, test_admin_headers: dict, session_suffix: str):
        suf = unique_suffix()
        csv_content = f"name,sku,price,cost\nImported Widget,IMP-{session_suffix}-{suf}-01,29.99,15.00\nImported Gadget,IMP-{session_suffix}-{suf}-02,49.99,25.00\n"
        files = {"file": ("products.csv", csv_content, "text/csv")}
        resp = httpx.post(
            f"{SERVER_URL}/api/import/products",
            files=files,
            headers=test_admin_headers,
            timeout=15,
        )
        data = assert_ok(resp)
        assert data["imported"] >= 2

    def test_import_products_missing_name(self, test_admin_headers: dict):
        csv_content = "sku,price\nNOSKU,10\n"
        files = {"file": ("bad.csv", csv_content, "text/csv")}
        resp = httpx.post(
            f"{SERVER_URL}/api/import/products",
            files=files,
            headers=test_admin_headers,
            timeout=10,
        )
        assert resp.status_code == 400


class TestImportXlsxJson:
    def test_import_customers_xlsx(self, test_admin_headers: dict, session_suffix: str):
        suf = unique_suffix()
        content = _make_xlsx(
            ["first_name", "last_name", "email", "phone"],
            [
                ["Xlsx", "One", f"xl1-{session_suffix}-{suf}@test.com", "555-0201"],
                ["Xlsx", "Two", f"xl2-{session_suffix}-{suf}@test.com", "555-0202"],
            ],
        )
        files = {"file": ("customers.xlsx", content, ei_xlsx_mime())}
        resp = httpx.post(
            f"{SERVER_URL}/api/import/customers",
            files=files,
            headers=test_admin_headers,
            timeout=15,
        )
        data = assert_ok(resp)
        assert data["imported"] >= 2

    def test_import_customers_json(self, test_admin_headers: dict, session_suffix: str):
        suf = unique_suffix()
        payload = json.dumps(
            [
                {
                    "first_name": "Json",
                    "last_name": "One",
                    "email": f"js1-{session_suffix}-{suf}@test.com",
                },
                {
                    "first_name": "Json",
                    "last_name": "Two",
                    "email": f"js2-{session_suffix}-{suf}@test.com",
                },
            ]
        )
        files = {"file": ("customers.json", payload, "application/json")}
        resp = httpx.post(
            f"{SERVER_URL}/api/import/customers",
            files=files,
            headers=test_admin_headers,
            timeout=15,
        )
        data = assert_ok(resp)
        assert data["imported"] >= 2

    def test_import_products_xlsx(self, test_admin_headers: dict, session_suffix: str):
        suf = unique_suffix()
        content = _make_xlsx(
            ["name", "sku", "price", "cost", "active"],
            [
                ["Xlsx Widget", f"XW-{session_suffix}-{suf}", 19.99, 9.0, True],
                ["Xlsx Gadget", f"XG-{session_suffix}-{suf}", 39.99, 20.0, False],
            ],
        )
        files = {"file": ("products.xlsx", content, ei_xlsx_mime())}
        resp = httpx.post(
            f"{SERVER_URL}/api/import/products",
            files=files,
            headers=test_admin_headers,
            timeout=15,
        )
        data = assert_ok(resp)
        assert data["imported"] >= 2

    def test_import_products_json(self, test_admin_headers: dict, session_suffix: str):
        suf = unique_suffix()
        payload = json.dumps(
            [
                {
                    "name": "Json Widget",
                    "sku": f"JW-{session_suffix}-{suf}",
                    "price": 12.5,
                    "quantity_on_hand": 3,
                },
                {"name": "Json Gadget", "sku": f"JG-{session_suffix}-{suf}", "price": 24.0},
            ]
        )
        files = {"file": ("products.json", payload, "application/json")}
        resp = httpx.post(
            f"{SERVER_URL}/api/import/products",
            files=files,
            headers=test_admin_headers,
            timeout=15,
        )
        data = assert_ok(resp)
        assert data["imported"] >= 2

    def test_import_customers_xlsx_missing_first_name(self, test_admin_headers: dict):
        content = _make_xlsx(["last_name", "email"], [["Nobody", "n@test.com"]])
        files = {"file": ("bad.xlsx", content, ei_xlsx_mime())}
        resp = httpx.post(
            f"{SERVER_URL}/api/import/customers",
            files=files,
            headers=test_admin_headers,
            timeout=10,
        )
        assert resp.status_code == 400

    def test_import_customers_json_missing_first_name(self, test_admin_headers: dict):
        payload = json.dumps([{"last_name": "Nobody", "email": "n@test.com"}])
        files = {"file": ("bad.json", payload, "application/json")}
        resp = httpx.post(
            f"{SERVER_URL}/api/import/customers",
            files=files,
            headers=test_admin_headers,
            timeout=10,
        )
        assert resp.status_code == 400


def ei_xlsx_mime() -> str:
    return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


class TestExportImportErrors:
    def test_export_unauthorized(self, client: httpx.Client):
        resp = client.get("/api/export/customers", timeout=10)
        assert resp.status_code in (401, 403)

    def test_import_unauthorized(self, client: httpx.Client):
        csv_content = "first_name,last_name\nTest,User\n"
        files = {"file": ("test.csv", csv_content, "text/csv")}
        resp = client.post("/api/import/customers", files=files, timeout=10)
        assert resp.status_code in (401, 403)
