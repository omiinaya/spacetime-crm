"""Unit tests for routes/export_import.py — CSV / XLSX / JSON export + import.

These run without a live STDB/CRM server: ``_sql`` and ``_call`` are
monkeypatched, and the route functions are invoked directly with a fake
user dict (the FastAPI Depends layer is bypassed).
"""

from __future__ import annotations

import io
import json

import pytest
from fastapi import HTTPException, UploadFile
from routes import export_import as ei

USER = {"id": "u_1", "tenant_id": "t_1", "name": "admin", "role": "admin"}

CUSTOMER_ROW = {
    "id": "c_1",
    "tenant_id": "t_1",
    "first_name": "Alice",
    "last_name": "Smith",
    "email": "alice@test.com",
    "phone": "555-0101",
    "portal_password_hash": "should-be-stripped",
}
PRODUCT_ROW = {
    "id": "p_1",
    "tenant_id": "t_1",
    "name": "Widget",
    "sku": "W-1",
    "price": 9.99,
    "cost": 4.5,
    "quantity_on_hand": 10,
    "active": True,
}


def _make_xlsx(headers: list[str], rows: list[list]) -> bytes:
    """Build an XLSX byte payload from headers + row values."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class FakeSql:
    """Async _sql stand-in that returns canned rows and records the query."""

    def __init__(self, rows: list[dict] | None = None) -> None:
        self.rows = rows or []
        self.queries: list[str] = []

    async def __call__(self, query: str, *args, **kwargs) -> list[dict]:
        self.queries.append(query)
        return self.rows


class FakeCall:
    """Async _call stand-in that records (reducer, args) tuples."""

    def __init__(self) -> None:
        self.calls: list[tuple[str, list]] = []

    async def __call__(self, reducer: str, args: list) -> None:
        self.calls.append((reducer, args))


# ── Format detection (pure helpers) ──


class TestFormatDetection:
    def test_csv_detection(self):
        rows, fieldnames, fmt = ei._payload_rows(b"first_name,last_name\nAlice,Smith\n")
        assert fmt == "csv"
        assert fieldnames == ["first_name", "last_name"]
        assert rows[0]["last_name"] == "Smith"

    def test_csv_with_bom_detection(self):
        rows, fieldnames, fmt = ei._payload_rows(b"\xef\xbb\xbffirst_name,last_name\nAlice,Smith\n")
        assert fmt == "csv"
        assert fieldnames == ["first_name", "last_name"]

    def test_json_detection(self):
        rows, fieldnames, fmt = ei._payload_rows(
            b'[{"first_name": "Bob", "last_name": "Jones", "active": true}]'
        )
        assert fmt == "json"
        assert fieldnames == ["first_name", "last_name", "active"]
        assert rows[0]["active"] is True

    def test_json_non_array_rejected(self):
        with pytest.raises(HTTPException) as exc:
            ei._payload_rows(b'{"customers": []}')
        assert "array of objects" in str(exc.value.detail)

    def test_json_array_of_non_objects_rejected(self):
        with pytest.raises(HTTPException) as exc:
            ei._payload_rows(b'["a", "b"]')
        assert "array of objects" in str(exc.value.detail)

    def test_xlsx_detection_and_round_trip(self):
        content = _make_xlsx(
            ["first_name", "last_name", "active"],
            [["Carol", "Davis", True], ["Eve", "Frank", False]],
        )
        assert content[:4] == b"PK\x03\x04"
        rows, fieldnames, fmt = ei._payload_rows(content)
        assert fmt == "xlsx"
        assert fieldnames == ["first_name", "last_name", "active"]
        assert rows[0]["first_name"] == "Carol"
        assert rows[0]["active"] is True
        assert rows[1]["active"] is False

    def test_cell_helpers(self):
        assert ei._cell_str(None) == ""
        assert ei._cell_str(False) == "false"
        assert ei._cell_str(9.99) == "9.99"
        assert ei._cell_default(None, "x") == "x"
        assert ei._cell_default(5, "x") == 5
        assert ei._parse_bool(True) is True
        assert ei._parse_bool("yes") is True
        assert ei._parse_bool(0) is False
        assert ei._parse_bool("false") is False


# ── Export ──


class TestExport:
    async def test_export_csv_default_format(self, monkeypatch):
        fake = FakeSql([CUSTOMER_ROW])
        monkeypatch.setattr(ei, "_sql", fake)
        resp = await ei.export_entity("customers", "csv", USER)
        assert resp.media_type == "text/csv"
        body = resp.body.decode()
        assert "first_name" in body and "Alice" in body
        assert "portal_password_hash" not in body  # sensitive field stripped
        assert fake.queries and "t_1" in fake.queries[0]  # tenant scoped

    async def test_export_json(self, monkeypatch):
        fake = FakeSql([CUSTOMER_ROW])
        monkeypatch.setattr(ei, "_sql", fake)
        resp = await ei.export_entity("customers", "json", USER)
        assert resp.media_type == "application/json"
        data = json.loads(resp.body)
        assert data[0]["first_name"] == "Alice"
        assert "portal_password_hash" not in data[0]
        assert "attachment" in resp.headers["content-disposition"]
        assert "customers.json" in resp.headers["content-disposition"]

    async def test_export_xlsx_round_trip(self, monkeypatch):
        fake = FakeSql([CUSTOMER_ROW, {**CUSTOMER_ROW, "id": "c_2", "first_name": "Bob"}])
        monkeypatch.setattr(ei, "_sql", fake)
        resp = await ei.export_entity("customers", "xlsx", USER)
        assert resp.media_type == ei.XLSX_MIME
        assert resp.body[:4] == b"PK\x03\x04"
        # Parse the exported workbook back with openpyxl
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(resp.body), read_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        headers = [str(h) for h in next(it)]
        assert "first_name" in headers
        assert "portal_password_hash" not in headers
        values = [next(it), next(it)]
        names = {v[headers.index("first_name")] for v in values}
        assert names == {"Alice", "Bob"}

    async def test_export_xlsx_empty_rows(self, monkeypatch):
        fake = FakeSql([])
        monkeypatch.setattr(ei, "_sql", fake)
        resp = await ei.export_entity("products", "xlsx", USER)
        assert resp.media_type == ei.XLSX_MIME
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(resp.body), read_only=True)
        assert wb.active.max_row in (0, 1)

    async def test_export_unknown_entity(self, monkeypatch):
        fake = FakeSql([])
        monkeypatch.setattr(ei, "_sql", fake)
        with pytest.raises(HTTPException) as exc:
            await ei.export_entity("nonexistent", "csv", USER)
        assert exc.value.status_code == 400

    async def test_export_bad_format(self, monkeypatch):
        fake = FakeSql([])
        monkeypatch.setattr(ei, "_sql", fake)
        with pytest.raises(HTTPException) as exc:
            await ei.export_entity("customers", "pdf", USER)
        assert exc.value.status_code == 400
        assert "csv, xlsx, json" in str(exc.value.detail)


# ── Import ──


class TestImportCustomers:
    async def test_import_customers_from_xlsx(self, monkeypatch):
        rec = FakeCall()
        monkeypatch.setattr(ei, "_call", rec)
        content = _make_xlsx(
            ["first_name", "last_name", "email", "phone"],
            [
                ["Imp", "One", "imp1@test.com", "555-0101"],
                ["Imp", "Two", "imp2@test.com", "555-0102"],
            ],
        )
        file = UploadFile(file=io.BytesIO(content), filename="customers.xlsx")
        result = await ei.import_customers(file=file, user=USER)
        assert result["imported"] == 2
        assert result["errors"] == []
        reducers = [c[0] for c in rec.calls]
        assert reducers == ["create_customer", "create_customer"]
        assert rec.calls[0][1] == ["t_1", "Imp", "One", "imp1@test.com", "555-0101"]

    async def test_import_customers_from_json(self, monkeypatch):
        rec = FakeCall()
        monkeypatch.setattr(ei, "_call", rec)
        payload = json.dumps(
            [{"first_name": "Imp", "last_name": "One", "email": "j1@test.com"}]
        ).encode()
        file = UploadFile(file=io.BytesIO(payload), filename="customers.json")
        result = await ei.import_customers(file=file, user=USER)
        assert result["imported"] == 1
        assert rec.calls[0][1] == ["t_1", "Imp", "One", "j1@test.com", ""]

    async def test_import_customers_accepts_export_spelling_of_address(self, monkeypatch):
        """Exported files use address_line_1/2 (DB columns); legacy CSV used
        address_line1/2. Both must import into the same reducer slots."""
        rec = FakeCall()
        monkeypatch.setattr(ei, "_call", rec)
        content = _make_xlsx(
            ["id", "first_name", "last_name", "address_line_1", "address_line_2"],
            [["c_addr1", "Addr", "Person", "123 Main St", "Apt 4B"]],
        )
        file = UploadFile(file=io.BytesIO(content), filename="customers.xlsx")
        result = await ei.import_customers(file=file, user=USER)
        assert result["imported"] == 1
        reducer, args = rec.calls[0]
        assert reducer == "import_customer"
        assert args[7] == "123 Main St" and args[8] == "Apt 4B"

    async def test_import_customers_with_id_preserves_ids(self, monkeypatch):
        rec = FakeCall()
        monkeypatch.setattr(ei, "_call", rec)
        content = _make_xlsx(
            ["id", "first_name", "last_name", "created_at", "updated_at"],
            [["c_99", "Preserve", "Me", 1720000000000, 1720000000000]],
        )
        file = UploadFile(file=io.BytesIO(content), filename="customers.xlsx")
        result = await ei.import_customers(file=file, user=USER)
        assert result["imported"] == 1
        reducer, args = rec.calls[0]
        assert reducer == "import_customer"
        assert args[1] == "c_99" and args[2] == "Preserve" and args[15] == 1720000000000

    async def test_import_customers_xlsx_missing_first_name(self, monkeypatch):
        monkeypatch.setattr(ei, "_call", FakeCall())
        content = _make_xlsx(["last_name", "email"], [["Nobody", "n@test.com"]])
        file = UploadFile(file=io.BytesIO(content), filename="bad.xlsx")
        with pytest.raises(HTTPException) as exc:
            await ei.import_customers(file=file, user=USER)
        assert exc.value.status_code == 400
        assert "first_name" in str(exc.value.detail)

    async def test_import_customers_json_missing_first_name(self, monkeypatch):
        monkeypatch.setattr(ei, "_call", FakeCall())
        payload = b'[{"last_name": "Nobody", "email": "n@test.com"}]'
        file = UploadFile(file=io.BytesIO(payload), filename="bad.json")
        with pytest.raises(HTTPException) as exc:
            await ei.import_customers(file=file, user=USER)
        assert exc.value.status_code == 400
        assert "first_name" in str(exc.value.detail)


class TestImportProducts:
    async def test_import_products_from_xlsx_with_active_bools(self, monkeypatch):
        rec = FakeCall()
        monkeypatch.setattr(ei, "_call", rec)
        content = _make_xlsx(
            ["name", "sku", "price", "cost", "active"],
            [["Widget", "W-1", 29.99, 15.0, True], ["Gadget", "G-1", 9.99, 4.0, False]],
        )
        file = UploadFile(file=io.BytesIO(content), filename="products.xlsx")
        result = await ei.import_products(file=file, user=USER)
        assert result["imported"] == 2
        assert result["errors"] == []
        reducer, args = rec.calls[0]
        assert reducer == "create_product"
        assert args[1] == "Widget" and args[2] == "W-1" and args[6] == 29.99
        # active not part of create_product args — the bools must not crash parsing
        reducer2, args2 = rec.calls[1]
        assert args2[1] == "Gadget"

    async def test_import_products_from_json(self, monkeypatch):
        rec = FakeCall()
        monkeypatch.setattr(ei, "_call", rec)
        payload = b'[{"name": "Widget", "sku": "W-1", "price": 29.99, "quantity_on_hand": 5}]'
        file = UploadFile(file=io.BytesIO(payload), filename="products.json")
        result = await ei.import_products(file=file, user=USER)
        assert result["imported"] == 1
        reducer, args = rec.calls[0]
        assert reducer == "create_product"
        assert args[1] == "Widget" and args[6] == 29.99 and args[8] == 5.0

    async def test_import_products_with_id(self, monkeypatch):
        rec = FakeCall()
        monkeypatch.setattr(ei, "_call", rec)
        content = _make_xlsx(
            ["id", "name", "sku", "created_at"],
            [["p_99", "Widget", "W-9", 1720000000000]],
        )
        file = UploadFile(file=io.BytesIO(content), filename="products.xlsx")
        result = await ei.import_products(file=file, user=USER)
        assert result["imported"] == 1
        reducer, args = rec.calls[0]
        assert reducer == "import_product"
        assert args[1] == "p_99" and args[2] == "Widget" and args[15] == 1720000000000

    async def test_import_products_missing_name(self, monkeypatch):
        monkeypatch.setattr(ei, "_call", FakeCall())
        payload = b'[{"sku": "NOSKU", "price": 10}]'
        file = UploadFile(file=io.BytesIO(payload), filename="bad.json")
        with pytest.raises(HTTPException) as exc:
            await ei.import_products(file=file, user=USER)
        assert exc.value.status_code == 400
        assert "name" in str(exc.value.detail)

    async def test_import_products_collects_row_errors(self, monkeypatch):
        rec = FakeCall()
        monkeypatch.setattr(ei, "_call", rec)
        # Second row has a non-numeric price → row error, first row still imported
        content = _make_xlsx(
            ["name", "price"],
            [["Good", "10.00"], ["Bad", "not-a-number"]],
        )
        file = UploadFile(file=io.BytesIO(content), filename="products.xlsx")
        result = await ei.import_products(file=file, user=USER)
        assert result["imported"] == 1
        assert len(result["errors"]) == 1
        assert "Row 3" in result["errors"][0]
